import os
from openpyxl import load_workbook
from itertools import zip_longest
from reportlab.lib.pagesizes import A4, portrait
from reportlab.pdfgen import canvas


def convert_to_pdf(photos_files, target_pdf_name):
    (w, h) = portrait(A4)
    c = canvas.Canvas(target_pdf_name, pagesize=portrait(A4))
    for photo_file_index in range(0, len(photos_files)):
        c.drawImage(os.getcwd() + '/jpgs/' + photos_files[photo_file_index], 0, 0, w, h, preserveAspectRatio=True)
        c.showPage()
    c.save()


def is_jpg(file_name):
    if 'jpg' in file_name.lower():
        if file_name != '0.jpg':
            return file_name


def grouper(iterable, n, fillvalue=None):
    # Collect data into fixed-length chunks or blocks
    # grouper('ABCDEFG', 3, 'x') --> ABC DEF Gxx"
    args = [iter(iterable)] * n
    return zip_longest(*args, fillvalue=fillvalue)


def get_corporation_list():
    l = []
    try:
        wb = load_workbook("扫描企业清单.xlsx")
        ws = wb.worksheets[0]
        rows = ws[1:ws.max_row]
        for row in rows:
            if row[0].value:
                l.append(row[0].value)
        return l, wb
    except FileNotFoundError:
        print('请将扫描企业名单表格命名为“扫描企业清单.xlsx”并放在程序同一目录中')


corp_list, loaded_wb = get_corporation_list()
print(corp_list)
all_files = os.listdir(os.getcwd() + '/jpgs/')
all_files.sort(key=lambda x: int(x[:-4]))
jpg_files = list(filter(is_jpg, all_files))
try:
    if len(jpg_files) % 5 != 0:
        raise ValueError
except ValueError:
    print("图片文件总数(" + str(len(jpg_files)) + "张)不是5的位数，请检查是否缺少或有多余图片。\n" 
                                            "每户必须有5张图片，不足请拍白纸占位\n"
                                            "文件请放在'jpgs'文件夹中，并且按顺序命名为数字")
    exit(0)

# for jpg in jpg_files:
#     print(jpg)
# convert_to_pdf(files)

print(list(grouper(jpg_files, 5)))
i = 0
for photo_pack in list(grouper(jpg_files, 5, '0.jpg')):
    try:
        if i < 9:
            pdf_name = '0' + str(i + 1) + '-' + corp_list[i] + '.pdf'
        else:
            pdf_name = str(i + 1) + '-' + corp_list[i] + '.pdf'
    except IndexError:
        # print(i)
        print("你只提供了%d户企业的名单，因此只将前%d张照片进行了合成。" % (i, i * 5))
        loaded_wb.save("合并结果.xlsx")
        exit(0)
    convert_to_pdf(photo_pack, pdf_name)
    for j in range(2, 2 + len(photo_pack)):
        loaded_wb.worksheets[0].cell(row=i + 1, column=j).value = photo_pack[j - 2]
        print(str(i + 1) + '-' + loaded_wb.worksheets[0].cell(i + 1, 1).value + ':' + photo_pack[j - 2])
    i += 1
loaded_wb.save("合并结果.xlsx")
