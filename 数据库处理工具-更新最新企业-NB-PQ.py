# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from string import Template
from datetime import date
import tkinter as tk  # 用于打开文件窗口
from tkinter import filedialog  # 用于打开文件窗口
import os


class Data(tk.Frame):
    def __init__(self, root):
        tk.Frame.__init__(self, root)
        root.withdraw()  # 隐藏空白的TK画板。
        self.content_type = ''
        self.data_content = ''
        self.data_tpl = ''
        self.div = ''
        self.div_list = ['SL', 'TB', 'YH']
        self.file_path = ''
        self.file_folder = ''
        self.head_tpl = ''
        self.head = ''
        now = date.today()
        self.today = "%d-%d-%d" %(now.year, now.month, now.day -1)
        self.choose_file()

    def choose_file(self):
        self.file_path = filedialog.askopenfilename()
        self.file_folder = os.path.dirname(self.file_path)

    def div_select(self):

        print('请在下列名单中选择对应的监管所代码：\n'
              '1、SL 狮岭\n'
              '2、YH 裕华\n'
              '3、TB 炭步\n'
              '请输入两个英文字符代码(必须大写):')
        self.div = input()

        while self.div not in self.div_list:
            print('输入有误，请重新输入')
            print('请在下列名单中选择对应的监管所代码：\n'
                  '1、SL 狮岭\n'
                  '2、YH 裕华\n'
                  '3、TB 炭步\n'
                  '请输入两个英文字符代码(必须大写):')
            self.div = input()

        content_type_list = ['1', '2', '3']
        print('请选择数据来源的类型：\n'
              '1、普通年报导出表格，只更新企业信息，不更新年报、核查、电联信息\n'
              '2、在普通年报表格内容的基础上增加电联、核查信息的表格\n'
              # '3、包含定位等全面信息的表格\n'
              '3、专项行动首次导入表格\n'
              '请输入数字:')
        self.content_type = input()

        while self.content_type not in content_type_list:
            print('请选择数据来源的类型：\n'
                  '1、普通年报导出表格，只更新企业信息，不更新年报、核查、电联信息\n'
                  '2、在普通年报表格内容的基础上增加电联、核查信息的表格\n'
                  '3、专项行动首次导入表格\n'
                  # '3、包含定位等全面信息的表格\n'
                  '请输入数字:')
            self.content_type = input()

        if self.content_type == '1':
            self.head_tpl = Template('''
                    UPDATE "public"."corps" SET "is_active" = false WHERE "corporation_aic_division" = '${div}';
                    insert into "public"."corps" 
                    ("corporation_name", "registration_num", "address", "phone", \
            "represent_person", "contact_person", "contact_phone", \
            "is_active", "corporation_aic_division" ) VALUES
                    ''')
            self.data_tpl = Template("('${c}','${r}','${a}','${p}', \
            '${rp}', '${cp}', '${cph}', true, '${div}'),\n")

        elif self.content_type == '2':
            self.head_tpl = Template('''
                    UPDATE "public"."corps" SET "is_active" = false WHERE "division" = '${div}';
                    insert into "public"."corps" 
                    ("corporation_name", "registration_num", "address", "phone", \
            "represent_person", "contact_person", "contact_phone", \
            'nian_bao_status', 'inspection_status', \
            'phone_call_record', "is_active", "corporation_aic_division" ) VALUES
                    ''')
            self.data_tpl = Template("('${c}','${r}','${a}','${p}', \
            '${rp}', '${cp}', '${cph}', \
            '${nbs}', '${ins}', \
            '${phcal}', true, '${div}'),\n")

        elif self.content_type == '3':
            self.head = '''
            insert into "public"."special_actions" 
            ("sp_name", "sp_num", "sp_corp_id", 
            "corporation_name", "registration_num", \
            "predefined_name", "predefined_registration_num", \
            "sp_aic_division" ) VALUES
                    '''
            self.data_tpl = Template("('${sname}','${snum}','${scid}',\
            '${c}', '${r}', \
            '${pn}', '${pnum}', \
            '${div}'),\n")

    def load_workbook(self):
        print('正在读取XLSX文件中的用户名单......')
        try:
            wb = load_workbook(self.file_path)
            self.ws = wb.worksheets[0]
        except FileNotFoundError:
            print('无法打开' + self.file_path)
            exit(0)

    def process_to_sql(self):
        if self.content_type == '1' or self.content_type == '2':
            print('正在处理用户名单......')
            i = 1
            rows = self.ws[3:self.ws.max_row]
            # rows = self.ws[5000:6000]
            for row in rows:
                c = r = a = rp = cp = ''

                try:
                    r = ''.join(row[1].value.split())
                except AttributeError:
                    print('存在注册号为空的情况，请检查后重新运行。')

                try:
                    c = ''.join(row[0].value.split())
                except AttributeError:
                    c = '(' + str(r) + ')无字号'

                try:
                    a = ''.join(row[2].value.split())
                except AttributeError:
                    a = ''

                try:
                    p = ''.join(row[3].value.split())
                except AttributeError:
                    p = ''

                # 以下读取年报情况列，并根据年报情况处理导入的内容，对于未填报者，增加更新时间标识。
                # try:
                #     nb = ''.join(row[4].value.split())
                # except AttributeError:
                #     nb = row[4].value
                # if nb == '未填报':
                #     nb = '17年度：截至' + self.today + nb
                # else:
                #     nb = '17年度：' + nb
                # row[5] 填报时间
                # row[6] 管辖机关

                try:
                    rp = ''.join(row[7].value.split())
                except AttributeError:
                    rp = ''

                try:
                    cp = ''.join(row[8].value.split())
                except AttributeError:
                    cp = ''

                try:
                    cph = ''.join(row[9].value.split())
                except AttributeError:
                    cph = ''

                try:
                    ins = row[10].value
                except IndexError:
                    pass
                try:
                    phcal = row[11].value
                except IndexError:
                    pass
                assert c != ''
                assert r != ''

                if self.content_type == '1':
                    self.data_content += self.data_tpl.substitute(c=c, r=r, a=a, p=p, rp=rp, cp=cp, cph=cph, div=self.div)
                elif self.content_type == '2':
                    self.data_content += self.data_tpl.substitute(c=c, r=r, a=a, p=p, rp=rp,
                                                                  cp=cp, cph=cph, ins=ins,
                                                                  phcal=phcal, div=self.div)
                print(str(i) + ':正在处理：' + c)
                i += 1

        elif self.content_type == '3':
            print('正在处理用户名单......')
            i = 1
            rows = self.ws[2:self.ws.max_row]
            for row in rows:
                sname = snum = scid = c = r = pn = pnum = ''

                try:
                    sname = ''.join(row[0].value.split())
                except AttributeError:
                    print('存在行动名称为空的情况，请检查表格，程序结束。')
                    exit(0)

                try:
                    snum = ''.join(row[1].value.split())
                except AttributeError:
                    print('存在行动代号为空的情况，请检查表格，程序结束。')
                    exit(0)

                try:
                    scid = row[2].value
                except AttributeError:
                    print('存在企业行动内序号为空的情况，请检查表格，程序结束。' + str(i))
                    exit(0)

                try:
                    c = ''.join(row[3].value.split())
                except AttributeError:
                    c = '(' + str(r) + ')无字号'

                try:
                    r = ''.join(row[4].value.split())
                except AttributeError:
                    print('存在注册号为空的情况，请检查后重新运行。')
                    exit(0)

                try:
                    pn = row[5].value
                except IndexError:
                    pass
                try:
                    pnum = row[6].value
                except IndexError:
                    pass

                assert c != ''
                assert r != ''

                self.data_content += self.data_tpl.substitute(sname=sname, snum=snum, scid=scid, c=c, r=r,
                                                              pn=pn, pnum=pnum, div=self.div)

                print(str(i) + ':正在处理：' + c)
                i += 1
                # ${sname}','${snum}','${scid}',\
                #             '${c}', '${r}', \
                #             '${pn}', '${pnum}', \
                #             '${div}

    def save_file(self):
        if self.content_type == '1' or self.content_type == '2':
            print('开始保存文件')
            print('保存路径：' + self.file_folder + os.path.sep)
            f = open(self.file_folder + os.path.sep + self.today + self.div + '-最新企业（无年报信息）.sql', 'wb')
            f.write(self.head.encode('utf8'))
            f.write(self.data_content[:-2].encode('utf8'))
            f.write(b'''
            ON CONFLICT (registration_num) DO UPDATE SET 
            corporation_name = EXCLUDED.corporation_name,
            address=EXCLUDED.address,
            represent_person = EXCLUDED.represent_person,
            contact_person = EXCLUDED.contact_person,
            corporation_aic_division = EXCLUDED.corporation_aic_division,
            is_active = EXCLUDED.is_active,
            phone = EXCLUDED.phone,
            contact_phone = EXCLUDED.contact_phone;''')
            f.close()
            print('保存完成，程序结束。')
            exit(0)
        elif self.content_type == '3':
            print('开始保存文件')
            print('保存路径：' + self.file_folder + os.path.sep)
            f = open(self.file_folder + os.path.sep + self.today + self.div + '-专项行动导入表.sql', 'wb')
            f.write(self.head.encode('utf8'))
            f.write(self.data_content[:-2].encode('utf8'))
            f.close()
            print('保存完成，程序结束。')
            exit(0)



if __name__ == '__main__':
    tk_root = tk.Tk()
    data = Data(tk_root)

    data.div_select()
    data.load_workbook()
    data.process_to_sql()
    data.save_file()
    tk_root.mainloop()
