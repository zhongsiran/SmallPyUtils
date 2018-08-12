import os 
import os.path 
import shutil
import re
import time
import datetime
import re
from openpyxl import load_workbook
from openpyxl import Workbook


def photoDetect(Dir): 
    for file in os.listdir(Dir): #历遍图片文件
        if ('jpg' in file.lower() or 'jpeg'in file.lower()):
            return True
        else:
            return False
        

def workbook_write(workbook, corp_name, row_num):
    wb = workbook
    ws = wb.active
    cell_name = 'A' + str(row_num)
    # print(cell_name)
    ws[cell_name] = corp_name
    return wb


if __name__ == '__main__':
    existphotos = []
    nophotos = []
    workbook = Workbook()
    for singledir, subdirs, files in os.walk(os.getcwd()):
        # print(singledir)
        corpname = re.sub(r'.*\\', "", singledir) # 删除剩下企业名称
        # corpname = re.sub(r'.*-',"",corpname) # 删除剩下企业名称
        if photoDetect(singledir):
            existphotos.append(corpname)
        else:
            nophotos.append(corpname)
    # print('有照片的文件夹')
    print(len(existphotos))
    for index in range(len(existphotos)):
        # print(str(index + 1) + '、' + existphotos[index])
        corp_name = re.sub('\S*-+', "", existphotos[index])
        print(corp_name)
        workbook = workbook_write(workbook, corp_name, index + 1)
    # for index in range(len(nophotos)):
    #     print(str(index + 1) + '、' + nophotos[index])
    workbook.save('有照片列表.xlsx')
    # os.system('pause')



        
        



