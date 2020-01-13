# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter as tk  # 用于打开文件窗口
from tkinter import filedialog  # 用于打开文件窗口
import tkinter.messagebox
import os


class Data(tk.Frame):
    def __init__(self, root):
        tk.Frame.__init__(self, root)
        root.withdraw()  # 隐藏空白的TK画板。
        self.file_path = ''
        self.file_folder = ''
        tkinter.messagebox.showwarning('提示','1、只能处理XLSX文件(无法选择其他格式的文件)\n2、将会在被选择的XLSX文件所在目录下生成文件夹\n3、以XLSX表格的第一张表为数据源。\n  a、默认跳过首行\n  b、将所有列从左到右合并作为文件夹名称，不希望出现在文件夹名称的数据请删去\n  c、如果要用文书生成器生成文书，请确保企业名称是最右边的一列！！')
        self.choose_file()

    def choose_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[('xlsx','*.xlsx')])
        self.file_folder = os.path.dirname(self.file_path)

    def load_workbook(self):
        print('正在读取XLSX文件中的用户名单......')
        try:
            wb = load_workbook(self.file_path)
            self.ws = wb.worksheets[0]
        except FileNotFoundError:
            print('无法打开' + self.file_path)
            exit(0)

    def make_folder(self):
        print('正在处理用户名单......')
        os.chdir(self.file_folder)
        i = 1
        rows = self.ws[2:self.ws.max_row]
        for row in rows:
            new_folder_name = ''
            for column in row:
                if new_folder_name != '':
                    new_folder_name = new_folder_name + '-' + str(column.value)
                else:
                    new_folder_name = str(column.value)
            
            try:
                os.mkdir(new_folder_name)
                print(str(i) + ':正在创建：' + new_folder_name)
            except FileExistsError:
                pass
                print(str(i) + ':跳过（已经存在）：' + new_folder_name)
            
            i += 1


if __name__ == '__main__':
    tk_root = tk.Tk()
    data = Data(tk_root)

    data.load_workbook()
    data.make_folder()
    tk_root.mainloop()
